"""
Turn content_plan.xlsx into a queue of scheduled Trial Reels.

What it does:
  1. Reads every content row (video name + hook + caption).
     - Caption is taken from the 'fixed_by_claude_caption' column if present,
       otherwise the 'Caption' column.
  2. Locates each unique video file on disk, copies it into the repo's
     videos/ folder under a clean slug, and records its public raw URL
     (https://raw.githubusercontent.com/<owner>/<repo>/<branch>/videos/<slug>).
  3. INTERLEAVES the posts so no two consecutive reels use the same source
     visual (round-robin across the distinct video files) — better A/B read.
  4. Assigns post times: PER_DAY posts per day, spaced EVERY minutes apart,
     starting at --start.
  5. Writes the jobs into data/queue.json as Trial Reels (graduation MANUAL).

Nothing is published here — run_due.py does that when each time arrives.

Usage:
  python build_schedule.py \
      --repo-dir /path/to/publisher_automatic \
      --owner katerini007 --repo publisher_automatic --branch main \
      --start "2026-08-04 09:00" --per-day 25 --every 38 \
      [--dry-run]
"""

import argparse
import re
import shutil
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

import queue_store

SHEET = Path(__file__).parent / "content_plan.xlsx"

# Where to look for the source videos named in the sheet.
SOURCE_DIRS = [
    Path("/Volumes/KINGSTON/Claude/AGENTS/VIDEO-CREATION/video editor/Video editing/"
         "For myself/SMM Strategy/The AI KAT/batch upload/batch 31st of July"),
    Path("/Volumes/KINGSTON/Claude/AGENTS/VIDEO-CREATION/video editor/Video editing/"
         "For myself/SMM Strategy/The AI KAT/batch upload/batch 31st of July/output"),
]

HEADER_ROW = 4  # row 4 holds headers; data starts row 5


def slug(name: str) -> str:
    """Clean a video filename into a URL/repo-safe slug, preserving extension."""
    stem = Path(name).stem
    ext = Path(name).suffix.lower() or ".mp4"
    stem = re.sub(r"[^A-Za-z0-9]+", "_", stem).strip("_").lower()
    return f"{stem}{ext}"


def find_source(name: str) -> Path:
    base = Path(name).name
    for d in SOURCE_DIRS:
        cand = d / base
        if cand.exists():
            return cand
    # case-insensitive / extension-insensitive fallback
    target = base.lower()
    for d in SOURCE_DIRS:
        if not d.exists():
            continue
        for f in d.iterdir():
            if f.name.lower() == target:
                return f
    raise FileNotFoundError(f"Source video not found for '{name}' in {SOURCE_DIRS}")


def read_rows():
    wb = openpyxl.load_workbook(SHEET)
    ws = wb.active
    headers = [c.value for c in ws[HEADER_ROW]]

    def col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    ci_name = col("Video Name")
    ci_hook = col("Hook")
    ci_cap = col("fixed_by_claude_caption", "Caption")
    if ci_name is None or ci_cap is None:
        raise SystemExit(f"Sheet missing required columns. Found: {headers}")

    rows = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        name = r[ci_name]
        if not name:
            continue
        caption = (r[ci_cap] or "").strip() if ci_cap is not None else ""
        hook = (r[ci_hook] or "").strip() if ci_hook is not None else ""
        rows.append({"name": str(name).strip(), "hook": hook, "caption": caption})
    return rows


def family(name: str) -> str:
    """The source visual a render belongs to, ignoring the '- hookNN' suffix.
    e.g. 'hf_...124758... - hook03.mp4' and '... - hook07.mp4' share a family."""
    stem = Path(name).stem
    return re.sub(r"\s*-\s*hook\s*\d+\s*$", "", stem, flags=re.IGNORECASE).strip().lower()


def interleave(rows):
    """Round-robin across visual FAMILIES so consecutive posts use different
    source visuals (A -> B -> C -> A ...), better for isolating the hook."""
    groups = {}
    order = []
    for row in rows:
        key = family(row["name"])
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)
    result = []
    while any(groups[n] for n in order):
        for n in order:
            if groups[n]:
                result.append(groups[n].pop(0))
    return result


def assign_times(rows, start: datetime, per_day: int, every_min: int):
    interval = timedelta(minutes=every_min)
    for i, row in enumerate(rows):
        day = i // per_day
        slot = i % per_day
        row["scheduled_time"] = start + timedelta(days=day) + slot * interval
    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--repo-dir", required=True, help="Local path to the cloned publisher_automatic repo")
    ap.add_argument("--owner", required=True)
    ap.add_argument("--repo", required=True)
    ap.add_argument("--branch", default="main")
    ap.add_argument("--start", required=True, help='First post time, e.g. "2026-08-04 09:00"')
    ap.add_argument("--per-day", type=int, default=25)
    ap.add_argument("--every", type=int, default=38, help="Minutes between posts within a day")
    ap.add_argument("--graduation", default="MANUAL", choices=["MANUAL", "SS_PERFORMANCE"])
    ap.add_argument("--dry-run", action="store_true", help="Show the plan; copy no files, write no queue")
    args = ap.parse_args()

    start = datetime.strptime(args.start, "%Y-%m-%d %H:%M")
    rows = read_rows()
    rows = interleave(rows)
    rows = assign_times(rows, start, args.per_day, args.every)

    repo_dir = Path(args.repo_dir)
    videos_dir = repo_dir / "videos"
    raw_base = f"https://raw.githubusercontent.com/{args.owner}/{args.repo}/{args.branch}/videos"

    if not args.dry_run:
        videos_dir.mkdir(parents=True, exist_ok=True)

    # Resolve + (optionally) copy each unique video once.
    slug_cache = {}
    for row in rows:
        if row["name"] not in slug_cache:
            src = find_source(row["name"])
            s = slug(row["name"])
            slug_cache[row["name"]] = s
            if not args.dry_run:
                dst = videos_dir / s
                if not dst.exists():
                    shutil.copy2(src, dst)
        row["slug"] = slug_cache[row["name"]]
        row["video_url"] = f"{raw_base}/{row['slug']}"

    # Build jobs
    jobs = []
    for row in rows:
        st = row["scheduled_time"]
        job = {
            "id": queue_store.new_job_id(st),
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "scheduled_time": st.isoformat(timespec="seconds"),
            "media_type": "REELS",
            "video_path": None,
            "image_path": None,
            "video_url": row["video_url"],
            "image_url": None,
            "caption": row["caption"],
            "trial": True,
            "graduation_strategy": args.graduation,
            "status": "scheduled",
            "container_id": None,
            "media_id": None,
            "permalink": None,
            "error": None,
            "attempts": 0,
            "_hook": row["hook"][:60],
        }
        jobs.append(job)

    # Report
    print(f"{'#':>3}  {'time':16}  {'slug':28}  hook")
    for i, (row, job) in enumerate(zip(rows, jobs), 1):
        print(f"{i:>3}  {job['scheduled_time'][:16]:16}  {row['slug'][:28]:28}  {row['hook'][:40]}")
    missing_cap = sum(1 for j in jobs if not j["caption"])
    print(f"\n{len(jobs)} jobs | {len(slug_cache)} unique videos | "
          f"{args.per_day}/day every {args.every}min | trial={args.graduation}")
    if missing_cap:
        print(f"⚠️  {missing_cap} jobs have NO caption — fill them before going live.")

    if args.dry_run:
        print("\n[dry-run] no files copied, no queue written.")
        return

    for job in jobs:
        queue_store.add_job(job)
    print(f"\n✓ wrote {len(jobs)} jobs to data/queue.json")
    print(f"✓ copied {len(slug_cache)} videos to {videos_dir}")


if __name__ == "__main__":
    main()
